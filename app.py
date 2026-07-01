import streamlit as st
import pandas as pd
from datetime import datetime
import zipfile
from io import BytesIO
import traceback
import openpyxl  # Added to extract unique Team names directly on file upload

from excel_reader import read_excel
from word_reader import read_word 
from word_updater import update_word
from excel_updater import update_excel
from config import TEAM_NAME_COLUMN, SPRINT_CYCLE_COLUMN, ALLOCATION_COLUMN, EXCEPTIONAL_RATE_COLUMN, RESOURCE_START_ROW

st.set_page_config(
    page_title="Automated Field Updater Tool", 
    page_icon="💼",
    layout="centered"
)

# --- GLOBAL ENTERPRISE STYLING ---
st.markdown("""
    <style>
        .main .block-container {
            padding-top: 2rem;
            padding-bottom: 3rem;
            max-width: 850px;
        }
        .app-title {
            font-size: 2.4rem !important;
            font-weight: 700 !important;
            color: #1E293B;
            margin-bottom: 0.2rem;
            text-align: center;
        }
        .app-subtitle {
            font-size: 1.05rem !important;
            color: #64748B;
            margin-bottom: 2rem;
            text-align: center;
        }
        .step-container {
            display: flex;
            justify-content: space-between;
            background-color: #F1F5F9;
            padding: 0.75rem 1.5rem;
            border-radius: 8px;
            margin-bottom: 2.5rem;
            border: 1px solid #E2E8F0;
        }
        .step-item {
            font-weight: 600;
            font-size: 0.9rem;
            color: #94A3B8;
        }
        .step-active {
            color: #0F172A !important;
            border-bottom: 2px solid #2563EB;
        }
        .section-header {
            font-size: 1.2rem !important;
            font-weight: 600 !important;
            color: #1E293B;
            margin-bottom: 1.2rem !important;
            border-left: 4px solid #2563EB;
            padding-left: 0.6rem;
        }
    </style>
""", unsafe_allow_html=True)

if "step" not in st.session_state:
    st.session_state.step = "upload"  
if "word_template" not in st.session_state:
    st.session_state.word_template = None
if "excel_template" not in st.session_state:
    st.session_state.excel_template = None
if "zip_buffer" not in st.session_state:
    st.session_state.zip_buffer = None
if "raw_word_bytes" not in st.session_state:
    st.session_state.raw_word_bytes = None
if "raw_excel_bytes" not in st.session_state:
    st.session_state.raw_excel_bytes = None
if "resource_tracker" not in st.session_state:
    st.session_state.resource_tracker = {}
# Added safe placeholders for the file names
if "word_filename" not in st.session_state:
    st.session_state.word_filename = "Document.docx"
if "excel_filename" not in st.session_state:
    st.session_state.excel_filename = "Spreadsheet.xlsx"

def reset_application():
    st.session_state.step = "upload"
    st.session_state.word_template = None
    st.session_state.excel_template = None
    st.session_state.zip_buffer = None
    st.session_state.raw_word_bytes = None
    st.session_state.raw_excel_bytes = None
    st.session_state.resource_tracker = {}
    st.session_state.word_filename = "Document.docx"
    st.session_state.excel_filename = "Spreadsheet.xlsx"

def process_initial_templates():
    if not st.session_state.get("uploaded_word_file") or not st.session_state.get("uploaded_excel_file"):
        st.error("Please ensure both base template files are selected before initiating parsing.")
        return

    with st.spinner("⏳ Extracting corporate configuration profiles... Please hold."):
        try:
            word_file = st.session_state.uploaded_word_file
            excel_file = st.session_state.uploaded_excel_file

            if not hasattr(word_file, "name") or not word_file.name.lower().endswith(".docx"):
                st.error("Please upload a valid .docx Word file (file extension .docx).")
                return

            st.session_state.raw_word_bytes = word_file.getvalue()
            st.session_state.raw_excel_bytes = excel_file.getvalue()
            
            # CRITICAL FIX: Save the names immediately while the widgets are actively populated
            st.session_state.word_filename = word_file.name
            st.session_state.excel_filename = excel_file.name

            st.session_state.word_template = read_word(st.session_state.raw_word_bytes)
            st.session_state.excel_template = read_excel(st.session_state.raw_excel_bytes)
            
            # Extract unique Team names and values to seed our live state tracker
            wb = openpyxl.load_workbook(BytesIO(st.session_state.raw_excel_bytes), data_only=True)
            
            # Explicitly target the sheet named "Team Loading Sheet"
            target_sheet = "Team Loading Sheet"
            if target_sheet in wb.sheetnames:
                ws = wb[target_sheet]
            else:
                ws = wb.worksheets[0]
                
            row = RESOURCE_START_ROW
            tracker = {}
            while True:
                team_val = ws[f"{TEAM_NAME_COLUMN}{row}"].value
                if team_val is None or str(team_val).strip() == "":
                    break
                t_name = str(team_val).strip()
                
                # Fetch baseline values straight from the spreadsheet row
                s_val = ws[f"{SPRINT_CYCLE_COLUMN}{row}"].value
                a_val = ws[f"{ALLOCATION_COLUMN}{row}"].value
                r_val = ws[f"{EXCEPTIONAL_RATE_COLUMN}{row}"].value
                
                # Safeguard float parsing
                try: sprint_init = float(s_val) if s_val is not None else 8.0
                except: sprint_init = 8.0
                try: 
                    alloc_init = float(a_val) if a_val is not None else 1.0
                    if alloc_init <= 1.0: alloc_init = alloc_init * 100.0
                except: alloc_init = 100.0
                try: rate_init = float(r_val) if r_val is not None else 125.0
                except: rate_init = 125.0

                tracker[t_name] = {
                    "sprint_cycle": sprint_init,
                    "allocation": alloc_init,
                    "exceptional_rate": rate_init
                }
                row += 1
                
            st.session_state.resource_tracker = tracker
            st.session_state.step = "edit"
        except Exception as e:
            st.error(f"Error parsing templates: {e}")
            st.exception(traceback.format_exc())

st.markdown('<div class="app-title">Automated Field Updater Tool</div>', unsafe_allow_html=True)
st.markdown('<div class="app-subtitle">Cross-document data synchronization and management portal</div>', unsafe_allow_html=True)

step = st.session_state.step
st.markdown(f"""
    <div class="step-container">
        <span class="step-item {'step-active' if step=='upload' else ''}">1. Template Ingestion</span>
        <span class="step-item {'step-active' if step=='edit' else ''}">2. Parameter Alignment</span>
        <span class="step-item {'step-active' if step=='download_ready' else ''}">3. Export Bundle</span>
    </div>
""", unsafe_allow_html=True)

# --- STEP 1: FILE UPLOAD STATE ---
if st.session_state.step == "upload":
    with st.container(border=True):
        st.markdown('<div class="section-header">Template Asset Upload</div>', unsafe_allow_html=True)
        st.write("Provide the source documents to extract existing baseline variables.")

        col1, col2 = st.columns(2)
        with col1:
            st.file_uploader("Statement of Work Document (SOW .docx)", type=["docx"], key="uploaded_word_file")
        with col2:
            st.file_uploader("Team Loading Sheet (TLS .xlsx)", type=["xlsx"], key="uploaded_excel_file")

    st.write("") 

    if st.session_state.get("uploaded_word_file") and st.session_state.get("uploaded_excel_file"):
        st.button(
            "Analyze and Extract Data Components", 
            width="stretch", 
            type="primary", 
            on_click=process_initial_templates
        )

# --- STEP 2: UNIFIED EDITING & CRUD STATE ---
elif st.session_state.step == "edit":
    word_fields_data = st.session_state.word_template.get("fields", {})
    excel_template = st.session_state.excel_template

    def parse_date(date_str, default_val):
        if not date_str: return default_val
        for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
            try: return datetime.strptime(str(date_str).strip().split(" ")[0], fmt).date()
            except ValueError: continue
        return default_val

    default_start = parse_date(excel_template.get("Project Start Date") or word_fields_data.get("Project Start Date"), datetime(2026, 8, 15).date())
    default_end = parse_date(excel_template.get("Project End Date") or word_fields_data.get("Project End Date"), datetime(2026, 12, 31).date())
    default_rel_start = parse_date(word_fields_data.get("RELEASE Start Date"), datetime(2026, 8, 16).date())
    default_rel_end = parse_date(word_fields_data.get("RELEASE End Date"), datetime(2026, 8, 18).date())
    default_approved_on = parse_date(word_fields_data.get("SOW Approved on"), datetime(2026, 8, 25).date())
    
    try:
        raw_cost = excel_template.get("Total Cost") or word_fields_data.get("Total Cost") or 0
        default_cost = float(str(raw_cost).replace("€", "").replace(",", "").strip())
    except ValueError:
        default_cost = 0.0

    with st.container(border=True):
        st.markdown('<div class="section-header">Global Core Parameters</div>', unsafe_allow_html=True)
        
        c1, c2, c3 = st.columns(3)
        with c1:
            project_start = st.date_input("Project Start Date", value=default_start)
        with c2:
            project_end = st.date_input("Project End Date", value=default_end)
        with c3:
            project_cost_input = st.number_input("Total Cost (€)", value=default_cost, step=100.0, format="%.2f")

        c4, c5 = st.columns(2)
        with c4:
            rel_start = st.date_input("Release Start Date", value=default_rel_start)
        with c5:
            rel_end = st.date_input("Release End Date (End of Warranty)", value=default_rel_end)

        c6, c7 = st.columns(2)
        with c6:
            approved_by = st.text_input("SOW Approved By (Email ID)", value=word_fields_data.get("SOW Approved by", "Aniket Saha"))
        with c7:
            approved_on = st.date_input("SOW Approved On", value=default_approved_on)

    st.write("") 

    # --- UPDATED INDIVIDUAL TEAM MANAGEMENT LAYOUT ---
    with st.container(border=True):
        st.markdown('<div class="section-header">Excel Operational Settings by Team Resource</div>', unsafe_allow_html=True)
        
        teams_available = list(st.session_state.resource_tracker.keys())
        
        if not teams_available:
            st.warning("⚠️ No records found in the configured Resource Table section range.")
        else:
            selected_team = st.selectbox("Select Team Name", options=teams_available)
            
            # Load stored metrics for this explicit team choice
            current_team_data = st.session_state.resource_tracker[selected_team]
            
            c8, c9, c10 = st.columns(3)
            with c8:
                sprint_cycle_input = st.number_input(
                    f"No of Sprint Cycle", 
                    value=float(current_team_data["sprint_cycle"]), 
                    step=0.5, 
                    format="%.2f",
                    key=f"sprint_{selected_team}"
                )
            with c9:
                allocation_input = st.number_input(
                    f"% of Allocation", 
                    value=int(current_team_data["allocation"]), 
                    step=5, 
                    format="%d",
                    key=f"alloc_{selected_team}"
                )
            with c10:
                exceptional_rate_input = st.number_input(
                    f"Exceptional Rate (€)", 
                    value=float(current_team_data["exceptional_rate"]), 
                    step=5.0, 
                    format="%.2f",
                    key=f"rate_{selected_team}"
                )
                
            # Keep changes synced immediately to memory state maps
            st.session_state.resource_tracker[selected_team] = {
                "sprint_cycle": sprint_cycle_input,
                "allocation": allocation_input,
                "exceptional_rate": exceptional_rate_input
            }

    st.write("") 

    with st.container(border=True):
        st.markdown('<div class="section-header">Payment Milestones Table Lifecycle</div>', unsafe_allow_html=True)
        initial_milestones = st.session_state.word_template.get("milestones")
        
        if initial_milestones is None or initial_milestones.empty:
            initial_milestones = pd.DataFrame(columns=["name", "date", "monthly", "quality", "invoice"])
        else:
            for col in ["monthly", "quality", "invoice"]:
                if col in initial_milestones.columns:
                    cleaned_series = initial_milestones[col].astype(str).str.replace("€", "", regex=False)
                    cleaned_series = cleaned_series.str.replace(",", "", regex=False).str.strip()
                    initial_milestones[col] = pd.to_numeric(cleaned_series, errors="coerce").astype(float).fillna(0.0)

        edited_milestones_df = st.data_editor(
            initial_milestones,
            num_rows="dynamic",
            width="stretch",
            column_config={
                "name": st.column_config.TextColumn("Milestone Name", required=True),
                "date": st.column_config.TextColumn("Payment Date", required=True),
                "monthly": st.column_config.NumberColumn("Monthly 85% (€)", format="€ %.2f", min_value=0.0, step=0.01),
                "quality": st.column_config.NumberColumn("Quality 15% (€)", format="€ %.2f", min_value=0.0, step=0.01),
                "invoice": st.column_config.NumberColumn("Invoice Amount (€)", format="€ %.2f", min_value=0.0, step=0.01),
            }
        )
    
    st.markdown("---")
    
    if st.button("Generate and Harmonize Documents", width="stretch", type="primary"):
        with st.spinner("⚙️ Executing cross-platform variable injection mapping..."):
            try:
                word_fields = {
                    "Project Start Date": project_start.strftime("%d-%b-%Y"),
                    "Project End Date": project_end.strftime("%d-%b-%Y"),
                    "Total Cost": f"{project_cost_input:.2f}",  
                    "RELEASE Start Date": rel_start.strftime("%d-%b-%Y"),
                    "RELEASE End Date": rel_end.strftime("%d-%b-%Y"),
                    "SOW Approved by": approved_by,
                    "SOW Approved on": approved_on.strftime("%d-%b-%Y")
                }

                excel_fields = {
                    "Project Start Date": project_start,  
                    "Project End Date": project_end,      
                    "Total Cost": project_cost_input
                }

                # Construct a normalized dict mapped for the excel backend parser (allocations divided back to base decimals)
                final_excel_tracker_payload = {}
                for t, datasets in st.session_state.resource_tracker.items():
                    final_excel_tracker_payload[t] = {
                        "sprint_cycle": datasets["sprint_cycle"],
                        "allocation": datasets["allocation"] / 100.0,
                        "exceptional_rate": datasets["exceptional_rate"]
                    }

                updated_word = update_word(BytesIO(st.session_state.raw_word_bytes), word_fields, edited_milestones_df)
                updated_excel = update_excel(BytesIO(st.session_state.raw_excel_bytes), excel_fields, final_excel_tracker_payload)

                # FIX: Pulling file names safely from the stable string attributes initialized during step 1
                orig_word_name = st.session_state.word_filename
                orig_excel_name = st.session_state.excel_filename

                zip_io = BytesIO()
                with zipfile.ZipFile(zip_io, "w", zipfile.ZIP_DEFLATED) as zip_file:
                    zip_file.writestr(f"Updated_{orig_word_name}", updated_word.getvalue())
                    zip_file.writestr(f"Updated_{orig_excel_name}", updated_excel.getvalue())
                
                zip_io.seek(0)
                st.session_state.zip_buffer = zip_io.getvalue()
                st.session_state.step = "download_ready"
                st.rerun()

            except Exception as e:
                st.error(f"Synchronization runtime error: {e}")
                st.exception(traceback.format_exc())

# --- STEP 3: DOWNLOAD & COMPLETE RESET STATE ---
elif st.session_state.step == "download_ready":
    with st.container(border=True):
        st.write("")
        st.markdown('<div style="font-size: 3.5rem; text-align: center; margin-bottom: 1rem;">🎉</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-header" style="border: none; padding: 0; text-align: center; font-size: 1.5rem !important;">Synchronization Pipeline Completed</div>', unsafe_allow_html=True)
        st.markdown('<div style="text-align: center; color: #64748B; margin-bottom: 2rem;">Both assets have been calibrated and compiled into a secure deployment package.</div>', unsafe_allow_html=True)

        st.download_button(
            label="📥 Export Synced Asset Package (ZIP Bundle)",
            data=st.session_state.zip_buffer,
            file_name="Synced_Corporate_Documents.zip",
            mime="application/zip",
            width="stretch",
            on_click=reset_application 
        )
        
        st.write("")
        if st.button("Discard Package & Process New Templates", width="stretch"):
            reset_application()
            st.rerun()
