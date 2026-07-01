from io import BytesIO
from openpyxl import load_workbook

from config import (
    EXCEL_COMMON_FIELDS,
    TEAM_NAME_COLUMN,
    SPRINT_CYCLE_COLUMN,
    ALLOCATION_COLUMN,
    EXCEPTIONAL_RATE_COLUMN,
    RESOURCE_START_ROW
)

class ExcelUpdater:

    def __init__(self, file):
        file_bytes = file.read() if hasattr(file, "read") else file
        if isinstance(file_bytes, memoryview):
            file_bytes = file_bytes.tobytes()
        self.workbook = load_workbook(BytesIO(file_bytes))
        
        # Target the specific sheet by name, fall back to first sheet if missing
        target_sheet = "Team Loading Sheet"
        if target_sheet in self.workbook.sheetnames:
            self.sheet = self.workbook[target_sheet]
        else:
            self.sheet = self.workbook.worksheets[0]

    def update_common_fields(self, common_fields):
        for field, cell in EXCEL_COMMON_FIELDS.items():
            if field in common_fields:
                val = common_fields[field]
                if field == "Total Cost":
                    try:
                        self.sheet[cell] = float(str(val).replace("€", "").replace(",", "").strip())
                    except ValueError:
                        self.sheet[cell] = val
                    self.sheet[cell].number_format = '€* #,##0.00'
                else:
                    self.sheet[cell] = val

    def update_resource_table(self, resource_data_dict):
        row = RESOURCE_START_ROW
        while True:
            team_name_raw = self.sheet[f"{TEAM_NAME_COLUMN}{row}"].value
            if team_name_raw is None or str(team_name_raw).strip() == "":
                break
            
            team_name = str(team_name_raw).strip()
            
            if team_name in resource_data_dict:
                data = resource_data_dict[team_name]
                
                # Update Sprint Cycle safely as float
                self.sheet[f"{SPRINT_CYCLE_COLUMN}{row}"] = float(data["sprint_cycle"])
                
                # Update Allocation as integer percentage without losing gridlines
                alloc_cell = self.sheet[f"{ALLOCATION_COLUMN}{row}"]
                alloc_cell.value = float(data["allocation"])
                alloc_cell.number_format = '0%'
                
                # Update Exceptional Rate safely as float
                self.sheet[f"{EXCEPTIONAL_RATE_COLUMN}{row}"] = float(data["exceptional_rate"])
                
            row += 1

    def update(self, common_fields, resource_data_dict):
        self.update_common_fields(common_fields)
        self.update_resource_table(resource_data_dict)

    def get_file(self):
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

def update_excel(uploaded_file, common_fields, resource_data_dict):
    updater = ExcelUpdater(uploaded_file)
    updater.update(common_fields, resource_data_dict)
    return updater.get_file()
